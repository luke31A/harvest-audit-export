#!/usr/bin/env python3
"""
Harvest API -> Excel Audit Export
==================================
Prompts for Harvest credentials and a date range, pulls all time entries
via the Harvest API (handling pagination), computes audit metrics, and
writes a formatted .xlsx report with two tabs:
  - Audit Summary : KPIs, breakdowns by employee/project, flagged entries
  - Raw Data      : Full entry table, auto-filtered, audit columns included

Build to .exe:  see build.bat in the project root
"""

import os
import sys
import time
import datetime as _dt
from datetime import datetime
from zoneinfo import ZoneInfo
from concurrent.futures import ThreadPoolExecutor, as_completed

import re

import requests
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

HARVEST_BASE_URL = "https://api.harvestapp.com/v2"

# Submission deadline: Monday after the work week at 9:30 AM Central
CENTRAL = ZoneInfo("America/Chicago")
DEADLINE_HOUR   = 9
DEADLINE_MINUTE = 30


def _submission_deadline(work_date: pd.Timestamp) -> pd.Timestamp:
    """
    Return the submission deadline for a given work date as a UTC-naive Timestamp.

    Rule: entries must be submitted by 9:30 AM Central on the Monday
    immediately following the work week (Mon–Sun).

    weekday() returns 0 for Monday, so (7 - weekday()) always gives
    the number of days forward to the next Monday, including 7 days
    forward for entries that fall on a Monday (deadline = following Monday).
    """
    days_ahead   = 7 - work_date.weekday()          # 1 (Sun) … 7 (Mon)
    next_monday  = work_date + pd.Timedelta(days=days_ahead)
    deadline_ct  = _dt.datetime(
        next_monday.year, next_monday.month, next_monday.day,
        DEADLINE_HOUR, DEADLINE_MINUTE,
        tzinfo=CENTRAL,
    )
    # Convert to UTC then strip tz so it matches our tz-naive Created At column
    return pd.Timestamp(deadline_ct).tz_convert("UTC").tz_localize(None)

# Colours (hex, no leading #)
C_NAV   = "1F3864"   # dark navy  – main header bg
C_BLUE  = "2E75B6"   # accent blue
C_LBLUE = "D6E4F0"   # light blue section header
C_KPI   = "EBF3FB"   # KPI card bg
C_ALT   = "F5F9FD"   # alternating row tint
C_FLAG  = "FCE4D6"   # orange tint for flagged rows
C_RED   = "C00000"   # deep red for audit flag section header
C_WHITE = "FFFFFF"
C_GREY  = "595959"

BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


# ---------------------------------------------------------------------------
# Credential & date prompts
# ---------------------------------------------------------------------------

def prompt_credentials() -> tuple:
    """Prompt the user for their Harvest Personal Access Token and Account ID."""
    print("\n" + "=" * 55)
    print("  Harvest API Audit Export  |  Commit Consulting")
    print("=" * 55)
    print("\nYou need a Harvest Personal Access Token and Account ID.")
    print("Generate them at: https://id.getharvest.com/developers\n")
    token = input("  Personal Access Token : ").strip()
    account_id = input("  Account ID            : ").strip()
    return token, account_id


def prompt_date_range() -> tuple:
    """Prompt the user for a start and end date (YYYY-MM-DD)."""
    print()
    while True:
        try:
            start = input("  Start date (YYYY-MM-DD) : ").strip()
            end   = input("  End date   (YYYY-MM-DD) : ").strip()
            start_dt = datetime.strptime(start, "%Y-%m-%d").date()
            end_dt   = datetime.strptime(end,   "%Y-%m-%d").date()
            if end_dt < start_dt:
                print("  End date must be on or after start date – try again.\n")
                continue
            return start, end
        except ValueError:
            print("  Invalid format – please use YYYY-MM-DD.\n")


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

def _headers(token: str, account_id: str) -> dict:
    return {
        "Authorization":    f"Bearer {token}",
        "Harvest-Account-ID": account_id,
        "User-Agent":       "CommitConsulting-HarvestExport/1.0",
    }


def _fetch_page(endpoint: str, headers: dict, params: dict) -> requests.Response:
    """Fetch one page, retrying after a wait on HTTP 429."""
    while True:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=30)
        if resp.status_code != 429:
            return resp
        time.sleep(15)


def fetch_all(endpoint: str, resource_key: str, token: str,
              account_id: str, extra_params: dict = None) -> list:
    """
    Fetch all records from a paginated Harvest endpoint.

    Page 1 is fetched first to discover total_pages; remaining pages
    are then fetched in parallel (up to 5 concurrent requests).
    """
    headers     = _headers(token, account_id)
    base_params = {**(extra_params or {}), "per_page": 100}

    # ── Page 1 (discover total_pages) ────────────────────────────────────
    resp = _fetch_page(endpoint, headers, {**base_params, "page": 1})

    if resp.status_code == 401:
        print("\n  ERROR: Invalid credentials (401). Check your token and account ID.")
        sys.exit(1)
    if resp.status_code == 403:
        return []
    if resp.status_code != 200:
        print(f"\n  WARNING: API returned {resp.status_code} for {endpoint} – skipping.")
        return []

    data        = resp.json()
    records     = list(data.get(resource_key, []))
    total_pages = data.get("total_pages", 1)

    print(f"    Page 1/{total_pages} ({len(records)} records)...", end="\r", flush=True)

    if total_pages <= 1:
        print()
        return records

    # ── Pages 2..N in parallel ───────────────────────────────────────────
    page_results: dict[int, list] = {}
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_page = {
            executor.submit(_fetch_page, endpoint, headers, {**base_params, "page": p}): p
            for p in range(2, total_pages + 1)
        }
        for future in as_completed(future_to_page):
            page_num = future_to_page[future]
            r        = future.result()
            page_results[page_num] = (
                r.json().get(resource_key, []) if r.status_code == 200 else []
            )
            so_far = len(records) + sum(len(v) for v in page_results.values())
            print(f"    {len(page_results) + 1}/{total_pages} pages fetched ({so_far} records)...",
                  end="\r", flush=True)

    print()
    for p in range(2, total_pages + 1):
        records.extend(page_results.get(p, []))

    return records


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def fetch_time_entries(token: str, account_id: str,
                       from_date: str, to_date: str) -> list:
    print(f"\n  Fetching time entries ({from_date} to {to_date})...")
    entries = fetch_all(
        endpoint     = f"{HARVEST_BASE_URL}/time_entries",
        resource_key = "time_entries",
        token        = token,
        account_id   = account_id,
        extra_params = {"from": from_date, "to": to_date},
    )
    print(f"  -> {len(entries)} entries fetched.")
    return entries


# ---------------------------------------------------------------------------
# Data parsing & transformation
# ---------------------------------------------------------------------------

def _safe(val):
    """Return None instead of NaN for missing values."""
    try:
        return None if pd.isna(val) else val
    except (TypeError, ValueError):
        return val


def parse_entries(entries: list) -> pd.DataFrame:
    """Flatten raw Harvest time entry JSON into a DataFrame."""
    rows = []
    for e in entries:
        row = {
            # ── Identifiers ──────────────────────────────────────────────
            "Entry ID":            e.get("id"),

            # ── Employee ─────────────────────────────────────────────────
            "Employee ID":         e.get("user", {}).get("id"),
            "Employee Name":       e.get("user", {}).get("name"),

            # ── Client & Project ─────────────────────────────────────────
            "Client":              e.get("client",  {}).get("name"),
            "Project":             e.get("project", {}).get("name"),
            "Project Code":        e.get("project", {}).get("code"),

            # ── Task ─────────────────────────────────────────────────────
            "Task":                e.get("task", {}).get("name"),

            # ── Time ─────────────────────────────────────────────────────
            "Work Date":           e.get("spent_date"),
            "Hours":               e.get("hours"),
            "Rounded Hours":       e.get("rounded_hours"),
            "Notes":               e.get("notes"),

            # ── Billability ──────────────────────────────────────────────
            "Billable":            e.get("billable"),
            "Billable Rate":       e.get("billable_rate"),
            "Billable Amount":     e.get("billable_amount"),
            "Cost Rate":           e.get("cost_rate"),
            "Cost Amount":         e.get("cost_amount"),

            # ── Timer detail (not visible in UI) ─────────────────────────
            "Timer Started At":    e.get("timer_started_at"),
            "Started Time":        e.get("started_time"),
            "Ended Time":          e.get("ended_time"),
            "Is Running":          e.get("is_running"),

            # ── Lock / Approval state ─────────────────────────────────────
            "Is Locked":           e.get("is_locked"),
            "Locked Reason":       e.get("locked_reason"),
            "Is Closed":           e.get("is_closed"),
            "Is Billed":           e.get("is_billed"),
            "Budgeted":            e.get("budgeted"),

            # ── Audit timestamps ─────────────────────────────────────────
            "Created At":          e.get("created_at"),
            "Updated At":          e.get("updated_at"),

            # ── Invoice link ─────────────────────────────────────────────
            "Invoice ID":          (e.get("invoice") or {}).get("id"),
            "Invoice Number":      (e.get("invoice") or {}).get("number"),

            # ── External reference ────────────────────────────────────────
            "External Ref":        (e.get("external_reference") or {}).get("id"),

        }
        rows.append(row)

    df = pd.DataFrame(rows)

    # ── Type coercions ────────────────────────────────────────────────────
    df["Work Date"]  = pd.to_datetime(df["Work Date"])
    df["Created At"] = pd.to_datetime(df["Created At"], utc=True).dt.tz_localize(None)
    df["Updated At"] = pd.to_datetime(df["Updated At"], utc=True).dt.tz_localize(None)

    return df


EXCLUDED_CLIENT_NAMES = {"Commit Consulting"}

def _check_notes_client(notes, own_client: str, all_clients: list) -> str:
    """
    Inspect an entry's notes field for client name references.

    Returns one of:
      "Client Match"                 — own client name found in notes
      "Possible Wrong Client — X"    — a different client's name found in notes
      "No Client Name Mentioned"     — no client name found (or notes is empty)

    Matching uses whole-word regex search (word boundaries) to avoid false
    positives from client names embedded inside other words (e.g. "KForce"
    inside "workforce"). Clients are tested longest-first to prefer more
    specific matches (e.g. "Newfold Digital" over "Newfold"). Client names
    shorter than 3 characters are skipped to avoid false positives.

    Clients in EXCLUDED_CLIENT_NAMES are never matched (e.g. "Commit Consulting").
    """
    if not notes or (isinstance(notes, float) and pd.isna(notes)):
        return "No Client Name Mentioned"

    notes_lower = str(notes).lower().strip()
    if not notes_lower:
        return "No Client Name Mentioned"

    own_lower = own_client.lower() if isinstance(own_client, str) else ""

    # Check own client first (skip excluded clients)
    if (
        own_lower
        and len(own_lower) >= 3
        and own_client not in EXCLUDED_CLIENT_NAMES
        and re.search(r'\b' + re.escape(own_lower) + r'\b', notes_lower)
    ):
        return "Client Match"

    # Check other clients, longest name first to reduce partial-match ambiguity
    for client in sorted(all_clients, key=len, reverse=True):
        if client == own_client:
            continue
        client_lower = client.lower()
        if (
            len(client_lower) >= 3
            and re.search(r'\b' + re.escape(client_lower) + r'\b', notes_lower)
        ):
            return f"Possible Wrong Client — {client}"

    return "No Client Name Mentioned"


def add_audit_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Derive audit metric columns from raw timestamps."""

    # Submission deadline per entry (9:30 AM Central, Monday after the work week)
    df["Submission Deadline"] = df["Work Date"].apply(_submission_deadline)

    # Current UTC time (tz-naive, matching Created At column) for unsubmitted entry check
    now_utc = pd.Timestamp(datetime.now(_dt.timezone.utc)).tz_localize(None)

    # Late logic:
    #   - If Created At is present: late when created after the deadline
    #   - If Created At is missing (entry not yet submitted): late when today is past the deadline
    has_created = df["Created At"].notna()
    df["Late Submission"] = (
        (has_created  & (df["Created At"] > df["Submission Deadline"])) |
        (~has_created & (now_utc         > df["Submission Deadline"]))
    )

    # Hours past deadline — use now_utc as the submission time for unsubmitted entries
    effective_submitted = df["Created At"].fillna(now_utc)
    df["Hours Past Deadline"] = (
        (effective_submitted - df["Submission Deadline"]).dt.total_seconds() / 3600
    ).clip(lower=0).round(1)

    # Raw submission lag in calendar days (still useful for context)
    df["Submission Lag (Days)"] = (
        df["Created At"].dt.normalize() - df["Work Date"]
    ).dt.days

    # Edit lag: hours between creation and last update
    edit_delta = (df["Updated At"] - df["Created At"]).dt.total_seconds() / 3600
    df["Edit Lag (Hours)"] = edit_delta.round(2)

    # Was the entry edited after initial creation? (> 3 min threshold)
    df["Was Edited"] = df["Edit Lag (Hours)"] > 0.05

    # Client name mention check in notes (exclude Commit Consulting and similar internal clients)
    all_clients = [
        c for c in df["Client"].dropna().unique()
        if isinstance(c, str) and c not in EXCLUDED_CLIENT_NAMES
    ]
    df["Notes: Client Check"] = df.apply(
        lambda row: _check_notes_client(row["Notes"], row["Client"], all_clients),
        axis=1,
    )

    # Blank notes flag
    df["Blank Notes"] = df["Notes"].isna() | (df["Notes"].str.strip() == "")

    return df


# ---------------------------------------------------------------------------
# Summary / analytics
# ---------------------------------------------------------------------------

def build_summary(df: pd.DataFrame) -> dict:
    n = len(df)

    # ── KPIs ─────────────────────────────────────────────────────────────
    late_n           = int(df["Late Submission"].sum())
    billable_hours   = round(df.loc[df["Billable"] == True, "Hours"].sum(), 2)

    kpis = {
        "Total Entries":                n,
        "Total Hours":                  round(df["Hours"].sum(), 2),
        "Total Billable Hours":         billable_hours,
        "% Billable":                   round(df["Billable"].sum() / n * 100, 1) if n else 0,
        "Late Entries":                 late_n,
        "% Late Submissions":           round(late_n / n * 100, 1) if n else 0,
        "% Entries Edited":             round(df["Was Edited"].sum() / n * 100, 1) if n else 0,
        "% Locked / Approved":          round(df["Is Locked"].sum() / n * 100, 1) if n else 0,
    }

    # ── By employee ───────────────────────────────────────────────────────
    by_emp = df.groupby("Employee Name", as_index=False).agg(
        Entries             = ("Entry ID",               "count"),
        Total_Hours         = ("Hours",                  "sum"),
        Avg_Sub_Lag         = ("Submission Lag (Days)",  "mean"),
        Pct_Edited          = ("Was Edited",             "mean"),
        Avg_Edit_Lag_Hrs    = ("Edit Lag (Hours)",       "mean"),
        Late_Submissions    = ("Late Submission",        "sum"),
        Pct_Late            = ("Late Submission",        "mean"),
        Blank_Notes         = ("Blank Notes",            "sum"),
        Locked              = ("Is Locked",              "sum"),
    )
    by_emp.columns = [
        "Employee", "Entries", "Total Hours",
        "Avg Submission Lag (Days)", "% Edited",
        "Avg Edit Lag (Hours)", "Late Submissions", "% Late",
        "Blank Notes", "Locked Entries",
    ]
    by_emp["% Edited"]                  = (by_emp["% Edited"] * 100).round(1)
    by_emp["% Late"]                    = (by_emp["% Late"]   * 100).round(1)
    by_emp["Total Hours"]               = by_emp["Total Hours"].round(2)
    by_emp["Avg Submission Lag (Days)"] = by_emp["Avg Submission Lag (Days)"].round(1)
    by_emp["Avg Edit Lag (Hours)"]      = by_emp["Avg Edit Lag (Hours)"].round(2)
    by_emp["Blank Notes"]               = by_emp["Blank Notes"].astype(int)
    by_emp = by_emp.sort_values("% Late", ascending=False).reset_index(drop=True)

    # ── By project ────────────────────────────────────────────────────────
    by_proj = df.groupby(["Client", "Project"], as_index=False).agg(
        Entries          = ("Entry ID",              "count"),
        Total_Hours      = ("Hours",                 "sum"),
        Avg_Sub_Lag      = ("Submission Lag (Days)", "mean"),
        Pct_Edited       = ("Was Edited",            "mean"),
        Late_Submissions = ("Late Submission",       "sum"),
    )
    by_proj.columns = [
        "Client", "Project", "Entries", "Total Hours",
        "Avg Submission Lag (Days)", "% Edited", "Late Submissions",
    ]
    by_proj["% Edited"]                  = (by_proj["% Edited"] * 100).round(1)
    by_proj["Total Hours"]               = by_proj["Total Hours"].round(2)
    by_proj["Avg Submission Lag (Days)"] = by_proj["Avg Submission Lag (Days)"].round(1)
    by_proj = by_proj.sort_values("Total Hours", ascending=False).reset_index(drop=True)

    # ── Audit flags ───────────────────────────────────────────────────────
    # Flag: submitted more than 7 days late  OR  edited after being locked
    flag_mask = df["Late Submission"] | (df["Was Edited"] & df["Is Locked"])
    flag_cols = [
        "Entry ID", "Employee Name", "Client", "Project",
        "Work Date", "Hours", "Submission Deadline", "Hours Past Deadline",
        "Submission Lag (Days)", "Edit Lag (Hours)",
        "Was Edited", "Is Locked", "Locked Reason", "Notes",
    ]
    flags = df.loc[flag_mask, flag_cols].reset_index(drop=True)

    return {"kpis": kpis, "by_employee": by_emp, "by_project": by_proj, "flags": flags}


def detect_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    """
    Identify potential duplicate time entries and return them as a flat DataFrame.

    Two match types:
      Hours match — same Employee, Client, Project, Date & Hours
      Notes match — same Employee, Client, Project, Date & Notes (non-empty)

    Each group of related entries shares a Duplicate Group ID.
    An entry can appear in both groups if it satisfies both conditions.
    """
    display_cols = [
        "Duplicate Group", "Match Type", "Entry ID", "Employee Name",
        "Client", "Project", "Task", "Work Date", "Hours",
        "Notes: Client Check", "Notes", "Created At", "Is Locked",
    ]

    results  = []
    group_id = 1

    # ── Hours match: same Employee + Client + Project + Date + Hours ──────
    hours_keys = ["Employee Name", "Client", "Project", "Work Date", "Hours"]
    for _, group in df.groupby(hours_keys, dropna=False):
        if len(group) < 2:
            continue
        grp = group.copy()
        grp["Duplicate Group"] = group_id
        grp["Match Type"]      = "Hours Match — Same Employee, Client, Project, Date & Hours"
        results.append(grp)
        group_id += 1

    # ── Notes match: same Employee + Client + Project + Date + Notes ──────
    # Normalise notes before grouping; skip entries with blank/null notes
    notes_df = df.copy()
    notes_df["_notes_norm"] = (
        notes_df["Notes"]
        .fillna("")
        .str.strip()
        .str.lower()
    )
    notes_df = notes_df[notes_df["_notes_norm"] != ""]   # exclude blank notes

    notes_keys = ["Employee Name", "Client", "Project", "Work Date", "_notes_norm"]
    for _, group in notes_df.groupby(notes_keys, dropna=False):
        if len(group) < 2:
            continue
        # Pull the original rows (without the normalised column)
        grp = df.loc[group.index].copy()
        grp["Duplicate Group"] = group_id
        grp["Match Type"]      = "Notes Match — Same Employee, Client, Project, Date & Notes"
        results.append(grp)
        group_id += 1

    if not results:
        return pd.DataFrame(columns=display_cols)

    combined = pd.concat(results, ignore_index=True)
    combined = combined.sort_values(["Duplicate Group", "Work Date", "Employee Name"])

    # Return only columns that exist in the DataFrame
    keep = ["Duplicate Group", "Match Type"] + [
        c for c in display_cols[2:] if c in combined.columns
    ]
    return combined[keep].reset_index(drop=True)


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold=False, size=10, color="000000") -> Font:
    return Font(bold=bold, size=size, color=color)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_header(ws, row: int, ncols: int, bg: str = C_NAV, fg: str = C_WHITE):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = _font(bold=True, size=10, color=fg)
        cell.fill      = _fill(bg)
        cell.alignment = _align(h="center")
        cell.border    = BORDER
    ws.row_dimensions[row].height = 26


def write_df_block(ws, df: pd.DataFrame, start_row: int,
                   title: str, title_bg: str = C_BLUE) -> int:
    """
    Write a titled DataFrame block to a worksheet.
    Returns the next available row number after the block.
    """
    ncols = len(df.columns)

    # Title row
    ws.cell(row=start_row, column=1, value=title)
    for c in range(1, ncols + 1):
        ws.cell(row=start_row, column=c).fill      = _fill(title_bg)
        ws.cell(row=start_row, column=c).font      = _font(bold=True, size=11, color=C_WHITE)
        ws.cell(row=start_row, column=c).border    = BORDER
        ws.cell(row=start_row, column=c).alignment = _align(v="center")
    ws.row_dimensions[start_row].height = 22

    # Column header row
    for c, col in enumerate(df.columns, 1):
        ws.cell(row=start_row + 1, column=c, value=col)
    style_header(ws, start_row + 1, ncols, bg=C_BLUE)

    # Data rows
    for r_offset, row_data in enumerate(df.itertuples(index=False), 2):
        r = start_row + r_offset
        alt = r_offset % 2 == 0
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(val, float) and pd.isna(val):
                cell.value = None
            elif hasattr(val, "item"):          # numpy scalar
                cell.value = val.item()
            elif isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime()
            else:
                cell.value = val
            cell.fill      = _fill(C_ALT if alt else C_WHITE)
            cell.font      = _font(size=9)
            cell.alignment = _align(v="center")
            cell.border    = BORDER

    return start_row + 2 + len(df)   # next available row


# ---------------------------------------------------------------------------
# Sheet: Audit Summary
# ---------------------------------------------------------------------------

def write_summary_sheet(wb: Workbook, summary: dict,
                        from_date: str, to_date: str):
    ws = wb.create_sheet("Audit Summary")
    ws.sheet_view.showGridLines = False

    # ── Banner ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:L2")
    banner = ws["A1"]
    banner.value     = f"Harvest Audit Report  |  {from_date}  to  {to_date}"
    banner.font      = _font(bold=True, size=15, color=C_WHITE)
    banner.fill      = _fill(C_NAV)
    banner.alignment = _align(h="center", v="center")
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 6

    # ── KPI cards (two rows of 4) ─────────────────────────────────────────
    kpi_items = list(summary["kpis"].items())

    def kpi_card(label, value, row, col):
        # Label
        lc = ws.cell(row=row, column=col, value=label)
        lc.font      = _font(size=8, color=C_GREY)
        lc.fill      = _fill(C_KPI)
        lc.alignment = _align(h="center", v="bottom")
        lc.border    = BORDER
        # Value
        vc = ws.cell(row=row + 1, column=col, value=value)
        vc.font      = _font(bold=True, size=16, color=C_BLUE)
        vc.fill      = _fill(C_KPI)
        vc.alignment = _align(h="center", v="top")
        vc.border    = BORDER

    for i, (label, value) in enumerate(kpi_items[:4]):
        kpi_card(label, value, row=3, col=i + 1)
    for i, (label, value) in enumerate(kpi_items[4:8]):
        kpi_card(label, value, row=6, col=i + 1)

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 6
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 28
    ws.row_dimensions[8].height = 14

    # ── By employee ───────────────────────────────────────────────────────
    next_row = write_df_block(ws, summary["by_employee"], 9,
                              "Breakdown by Employee")
    next_row += 2

    # ── By project ────────────────────────────────────────────────────────
    next_row = write_df_block(ws, summary["by_project"], next_row,
                              "Breakdown by Client / Project")
    next_row += 2

    # ── Audit flags ───────────────────────────────────────────────────────
    if not summary["flags"].empty:
        next_row = write_df_block(
            ws, summary["flags"], next_row,
            f"Audit Flags  ({len(summary['flags'])} entries) — "
            "Late Submissions & Edits After Lock",
            title_bg=C_RED,
        )
        # Tint flag rows orange
        flag_data_start = next_row - len(summary["flags"])
        ncols = len(summary["flags"].columns)
        for r in range(flag_data_start, next_row):
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = _fill(C_FLAG)

    # ── Column widths ─────────────────────────────────────────────────────
    widths = [28, 10, 14, 26, 22, 14, 16, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Sheet: Raw Data
# ---------------------------------------------------------------------------

def write_raw_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False

    headers = list(df.columns)
    ncols   = len(headers)

    # Header row
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, ncols, bg=C_NAV)

    # Column width map
    col_widths = {
        "Entry ID": 12,          "Employee ID": 12,      "Employee Name": 24,
        "Client": 24,            "Project": 30,          "Project Code": 14,
        "Task": 24,              "Work Date": 13,        "Hours": 9,
        "Rounded Hours": 12,     "Notes": 45,            "Billable": 10,
        "Billable Rate": 13,     "Billable Amount": 16,  "Cost Rate": 11,
        "Cost Amount": 13,       "Timer Started At": 20, "Started Time": 14,
        "Ended Time": 14,        "Is Running": 11,       "Is Locked": 11,
        "Locked Reason": 22,     "Is Closed": 11,        "Is Billed": 11,
        "Budgeted": 10,          "Created At": 22,       "Updated At": 22,
        "Invoice ID": 12,        "Invoice Number": 16,   "External Ref": 16,
        "Submission Deadline": 22,   "Hours Past Deadline": 18,
        "Submission Lag (Days)": 20, "Edit Lag (Hours)": 17,
        "Was Edited": 12,            "Late Submission": 16,
        "Notes: Client Check": 28,
    }

    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = col_widths.get(h, 14)

    # Data rows
    for r_offset, row_data in enumerate(df.itertuples(index=False), 2):
        row = r_offset
        alt = r_offset % 2 == 0

        # Default fill; override for flagged rows
        is_flagged = (
            df.iloc[r_offset - 2]["Late Submission"]
            or (df.iloc[r_offset - 2]["Was Edited"] and df.iloc[r_offset - 2]["Is Locked"])
        )
        row_fill = _fill(C_FLAG if is_flagged else (C_ALT if alt else C_WHITE))

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c)
            if isinstance(val, float) and pd.isna(val):
                cell.value = None
            elif hasattr(val, "item"):
                cell.value = val.item()
            elif isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime()
            else:
                cell.value = val
            cell.fill      = row_fill
            cell.font      = _font(size=9)
            cell.alignment = _align(v="center")
            cell.border    = BORDER

    ws.freeze_panes  = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


# ---------------------------------------------------------------------------
# Sheet: Duplicate Entries
# ---------------------------------------------------------------------------

def write_duplicates_sheet(wb: Workbook, dupes: pd.DataFrame):
    ws = wb.create_sheet("Duplicate Entries")
    ws.sheet_view.showGridLines = False

    # ── Empty state ───────────────────────────────────────────────────────
    if dupes.empty:
        ws.merge_cells("A1:H3")
        cell = ws["A1"]
        cell.value     = "No potential duplicate entries found in this date range."
        cell.font      = _font(bold=True, size=12, color=C_BLUE)
        cell.fill      = _fill(C_KPI)
        cell.alignment = _align(h="center", v="center")
        ws.row_dimensions[1].height = 50
        return

    ncols      = len(dupes.columns)
    num_groups = dupes["Duplicate Group"].nunique()
    num_hours  = dupes[dupes["Match Type"].str.startswith("Hours")]["Duplicate Group"].nunique()
    num_notes  = dupes[dupes["Match Type"].str.startswith("Notes")]["Duplicate Group"].nunique()

    # ── Banner ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(ncols)}2")
    banner       = ws["A1"]
    banner.value = (
        f"Potential Duplicate Entries  |  {len(dupes)} entries across "
        f"{num_groups} groups  "
        f"({num_hours} hours match, {num_notes} notes match)"
    )
    banner.font      = _font(bold=True, size=13, color=C_WHITE)
    banner.fill      = _fill(C_NAV)
    banner.alignment = _align(h="center", v="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 6

    # ── Legend row ────────────────────────────────────────────────────────
    legend = (
        "Orange: same employee, client, project, date & hours.   "
        "Blue: same employee, client, project, date & notes (non-empty)."
    )
    ws.cell(row=3, column=1, value=legend).font = _font(size=9, color=C_GREY)
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 6

    # ── Column headers ────────────────────────────────────────────────────
    for c, col in enumerate(dupes.columns, 1):
        ws.cell(row=5, column=c, value=col)
    style_header(ws, 5, ncols, bg=C_NAV)

    # ── Data rows — colour-coded by group and match type ──────────────────
    # Exact groups alternate between two orange shades; possible between two blue shades
    exact_shades = ["FCE4D6", "F8CBAD"]
    poss_shades  = [C_ALT,    "DCE6F1"]
    group_colour = {}
    exact_idx, poss_idx = 0, 0

    for r_offset, row_data in enumerate(dupes.itertuples(index=False), 6):
        grp_id     = row_data[0]   # Duplicate Group (first column)
        match_type = row_data[1]   # Match Type (second column)

        if grp_id not in group_colour:
            if match_type.startswith("Hours"):
                group_colour[grp_id] = exact_shades[exact_idx % 2]
                exact_idx += 1
            else:
                group_colour[grp_id] = poss_shades[poss_idx % 2]
                poss_idx += 1

        fill_color = group_colour[grp_id]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_offset, column=c)
            if isinstance(val, float) and pd.isna(val):
                cell.value = None
            elif hasattr(val, "item"):
                cell.value = val.item()
            elif isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime()
            else:
                cell.value = val
            cell.fill      = _fill(fill_color)
            cell.font      = _font(size=9)
            cell.alignment = _align(v="center")
            cell.border    = BORDER

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = {
        "Duplicate Group": 14,  "Match Type": 50,        "Entry ID": 12,
        "Employee Name": 24,    "Client": 24,             "Project": 28,
        "Task": 22,             "Work Date": 13,          "Hours": 9,
        "Notes: Client Check": 28, "Notes": 45,           "Created At": 22,
        "Is Locked": 11,
    }
    for c, col in enumerate(dupes.columns, 1):
        ws.column_dimensions[get_column_letter(c)].width = col_widths.get(col, 14)

    ws.freeze_panes    = "A6"
    ws.auto_filter.ref = f"A5:{get_column_letter(ncols)}5"


# ---------------------------------------------------------------------------
# Sheet: Blank Notes
# ---------------------------------------------------------------------------

def write_blank_notes_sheet(wb: Workbook, df: pd.DataFrame):
    blank_df = df[df["Blank Notes"] == True].copy()

    ws = wb.create_sheet("Blank Notes")
    ws.sheet_view.showGridLines = False

    # ── Empty state ───────────────────────────────────────────────────────
    if blank_df.empty:
        ws.merge_cells("A1:H3")
        cell = ws["A1"]
        cell.value     = "No entries with blank notes found in this date range."
        cell.font      = _font(bold=True, size=12, color=C_BLUE)
        cell.fill      = _fill(C_KPI)
        cell.alignment = _align(h="center", v="center")
        ws.row_dimensions[1].height = 50
        return

    display_cols = [
        "Entry ID", "Employee Name", "Client", "Project", "Task",
        "Work Date", "Hours", "Billable", "Billable Rate",
        "Is Locked", "Late Submission", "Created At", "Updated At",
    ]
    keep = [c for c in display_cols if c in blank_df.columns]
    blank_df = blank_df[keep].sort_values(["Employee Name", "Work Date"]).reset_index(drop=True)

    ncols = len(blank_df.columns)

    # ── Banner ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(ncols)}2")
    banner       = ws["A1"]
    banner.value = f"Entries with Blank Notes  |  {len(blank_df)} entries"
    banner.font      = _font(bold=True, size=13, color=C_WHITE)
    banner.fill      = _fill(C_NAV)
    banner.alignment = _align(h="center", v="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 6

    # ── Column headers ────────────────────────────────────────────────────
    for c, col in enumerate(blank_df.columns, 1):
        ws.cell(row=3, column=c, value=col)
    style_header(ws, 3, ncols, bg=C_NAV)

    # ── Data rows ─────────────────────────────────────────────────────────
    for r_offset, row_data in enumerate(blank_df.itertuples(index=False), 4):
        alt = r_offset % 2 == 0
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_offset, column=c)
            if isinstance(val, float) and pd.isna(val):
                cell.value = None
            elif hasattr(val, "item"):
                cell.value = val.item()
            elif isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime()
            else:
                cell.value = val
            cell.fill      = _fill(C_ALT if alt else C_WHITE)
            cell.font      = _font(size=9)
            cell.alignment = _align(v="center")
            cell.border    = BORDER

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = {
        "Entry ID": 12,       "Employee Name": 24,  "Client": 24,
        "Project": 28,        "Task": 22,           "Work Date": 13,
        "Hours": 9,           "Billable": 10,       "Billable Rate": 13,
        "Is Locked": 11,      "Late Submission": 16, "Created At": 22,
        "Updated At": 22,
    }
    for c, col in enumerate(blank_df.columns, 1):
        ws.column_dimensions[get_column_letter(c)].width = col_widths.get(col, 14)

    ws.freeze_panes    = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}3"


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def save_workbook(wb: Workbook, from_date: str, to_date: str) -> str:
    # Place output folder next to the script (or .exe)
    base_dir   = os.path.dirname(os.path.abspath(
        sys.executable if getattr(sys, "frozen", False) else __file__
    ))
    output_dir = os.path.join(base_dir, "output")
    os.makedirs(output_dir, exist_ok=True)

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Harvest_Audit_{from_date}_to_{to_date}_{ts}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    token, account_id = prompt_credentials()
    from_date, to_date = prompt_date_range()

    entries = fetch_time_entries(token, account_id, from_date, to_date)

    if not entries:
        print("\n  No time entries found for that date range.")
        input("\n  Press Enter to exit.")
        return

    print("\n  Processing data...")
    df      = parse_entries(entries)
    df      = add_audit_columns(df)
    summary = build_summary(df)
    dupes   = detect_duplicates(df)

    if dupes.empty:
        print("  -> No potential duplicate entries found.")
    else:
        print(f"  -> {len(dupes)} entries flagged across "
              f"{dupes['Duplicate Group'].nunique()} duplicate groups.")

    print("  Building Excel report...")
    wb = Workbook()
    del wb[wb.sheetnames[0]]   # remove default blank sheet

    write_summary_sheet(wb, summary, from_date, to_date)
    write_duplicates_sheet(wb, dupes)
    write_blank_notes_sheet(wb, df)
    write_raw_sheet(wb, df)

    filepath = save_workbook(wb, from_date, to_date)

    print(f"\n  Report saved to:\n    {filepath}\n")
    input("  Press Enter to exit.")


if __name__ == "__main__":
    main()
